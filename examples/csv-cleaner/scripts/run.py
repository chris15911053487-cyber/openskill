"""csv-cleaner — clean a CSV and export it as an .xlsx workbook.

OpenSkill spawns this script with the following environment:

  OPENSKILL_INPUT_FILE  — JSON file with the request body
  OPENSKILL_OUTPUT_DIR  — directory to write produced files into
  PYTHONPATH            — pre-installed pandas / openpyxl / ...

The script reads input, applies the requested cleaning steps, and writes
exactly one .xlsx file under OPENSKILL_OUTPUT_DIR.

Exit codes:
  0  — success (file written)
  1  — script failure (full traceback on stderr; runner returns SCRIPT_FAILED)
"""

import io
import json
import os
import sys
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def main() -> int:
    input_file = os.environ.get("OPENSKILL_INPUT_FILE")
    output_dir = os.environ.get("OPENSKILL_OUTPUT_DIR")
    if not input_file or not output_dir:
        print("OPENSKILL_INPUT_FILE / OPENSKILL_OUTPUT_DIR not set", file=sys.stderr)
        return 1

    with open(input_file, "r", encoding="utf-8") as f:
        body = json.load(f)

    # The route layer wraps user payloads as `{ "input": <user json> }`.
    # Accept either shape so this script works from both the Run tab and
    # ad-hoc curl tests.
    payload = body.get("input", body) if isinstance(body, dict) else {}

    csv_text = payload.get("csv")
    if not isinstance(csv_text, str) or not csv_text.strip():
        print("`csv` is required and must be a non-empty string", file=sys.stderr)
        return 1

    filename = payload.get("filename") or "cleaned.xlsx"
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    sheet_name = payload.get("sheetName") or "cleaned"
    dedup_columns = payload.get("dedupColumns") or []
    trim = bool(payload.get("trim", True))
    drop_empty = bool(payload.get("dropEmptyRows", True))

    df = pd.read_csv(io.StringIO(csv_text), dtype=str, keep_default_na=False)
    # pandas keeps header whitespace by default; trim once unconditionally so
    # downstream column references are predictable.
    df.columns = [str(c).strip() for c in df.columns]

    if trim:
        for col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    if drop_empty:
        df = df[df.apply(lambda row: any(str(v).strip() for v in row), axis=1)]

    if dedup_columns:
        missing = [c for c in dedup_columns if c not in df.columns]
        if missing:
            print(f"dedup column(s) not present: {missing}", file=sys.stderr)
            return 1
        df = df.drop_duplicates(subset=dedup_columns, keep="first")

    # Reset index so the .xlsx doesn't include weird gaps after dedup.
    df = df.reset_index(drop=True)

    out_path = Path(output_dir) / filename
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # We let pandas write through openpyxl, then fix up formatting.
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        # Bold header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Auto-size columns based on max cell length (cap at 60 to avoid
        # absurd widths).
        for idx, col in enumerate(df.columns, start=1):
            longest = max((len(str(v)) for v in df[col]), default=0)
            width = min(60, max(len(str(col)), longest) + 2)
            ws.column_dimensions[get_column_letter(idx)].width = width

    return 0


if __name__ == "__main__":
    sys.exit(main())
